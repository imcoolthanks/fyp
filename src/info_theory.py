import numpy as np
import pandas as pd
import itertools

from openpyxl import load_workbook

from src import kraskov_jidt

SURROGATE_FILEPATH = "./results/surrogate_data.xlsx"

NUM_SURROGATES = 100
SIGNIFICANCE_LEVEL = 5

# =============== Surrogate Data Generation ==================

def surrogate(func, df, *args):
# def surrogate(num_bins, func, df, *args):
    actual_value = func(df, *args)

    sheet_name = f"{func.__name__}_over_time" if (df.index.max() - df.index.min()).days < 15 * 365 else func.__name__
    column_name = args_to_column_name(df, args)

    try:
        surrogate_df = pd.read_excel(SURROGATE_FILEPATH, sheet_name=sheet_name, parse_dates=True)
        random_values = surrogate_df[column_name].values
    except:
        random_values = generate_surrogate_data(func, df, column_name, *args)

    lower_bound = np.percentile(random_values, SIGNIFICANCE_LEVEL)
    upper_bound = np.percentile(random_values, 100 - SIGNIFICANCE_LEVEL)
    return actual_value - np.mean(random_values) if actual_value < lower_bound or actual_value > upper_bound else 0

def args_to_column_name(df, args, num_bins = None):
    if num_bins is None:
        return f"{df.index.min().strftime('%Y-%m-%d')}_{df.index.max().strftime('%Y-%m-%d')}_{'_'.join(list(df.columns))}_{'_'.join([str(arg) for arg in args])}"
    else:
        return f"{df.index.min().strftime('%Y-%m-%d')}_{df.index.max().strftime('%Y-%m-%d')}_numbins{num_bins}_{'_'.join(list(df.columns))}_{'_'.join([str(arg) for arg in args])}"

def generate_surrogate_data(func, df, column_name, *args):
    # Generate surrogates
    def shift_timeseries(timeseries):
        shift = np.random.randint(-500, 500)
        return np.roll(timeseries, shift)
    
    if (df.index.max() - df.index.min()).days < 15 * 365:
        num_surrogates = 25
        sheet_name = f"t_{func.__name__}"
    else:
        num_surrogates = NUM_SURROGATES
        sheet_name = func.__name__ 
    
    random_values = []
    for _ in range(num_surrogates):
        surrogate_timeseries = df.apply(lambda x: shift_timeseries(x))
        random_values.append(func(surrogate_timeseries, *args))

    try:
        surrogate_df = pd.read_excel(SURROGATE_FILEPATH, sheet_name=sheet_name, index_col=0, parse_dates=True)
    except:
        workbook = load_workbook(SURROGATE_FILEPATH)
        workbook.create_sheet(sheet_name)
        workbook.save(SURROGATE_FILEPATH)
        surrogate_df = pd.DataFrame()

    surrogate_df[column_name] = random_values
    with pd.ExcelWriter(SURROGATE_FILEPATH, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        surrogate_df.to_excel(writer, sheet_name=sheet_name)

    return random_values

# ================= Mutual Information and Entropy ======================

def entropy(X):
    _,num_counts = np.unique(X,return_counts=True,axis=0)
    p = num_counts / sum(num_counts)
    ent = - np.sum(p * np.log(p))
    return ent

def joint_entropy(df):
    asset_timeseries = df_to_tuple_dictionary(df)
    return entropy(np.column_stack(list(asset_timeseries.values())))

def MI_timeseries(timeseries1, timeseries2):
    mi = entropy(timeseries1) + entropy(timeseries2) - entropy(np.column_stack((timeseries1, timeseries2)))
    return mi

def MI(df, asset1, asset2):
    X_series = df[asset1]
    Y_series = df[asset2]

    X = [tuple_if_list(x) for x in X_series.values] if isinstance(X_series, pd.DataFrame) or isinstance(X_series, pd.Series) else X_series
    Y = [tuple_if_list(x) for x in Y_series.values] if isinstance(Y_series, pd.DataFrame) or isinstance(Y_series, pd.Series) else Y_series
    mi = entropy(X) + entropy(Y) - entropy(np.column_stack((X,Y)))
    return mi

# ==================== Normalized Variability of Information ====================

def NVI(P1, P2):
    return (entropy(P1) + entropy(P2) - 2 * MI_timeseries(P1, P2)) / (entropy(P1) + entropy(P2) - MI_timeseries(P1, P2))

# ==================== Oinfo ======================
def Oinfo(df, continuous = False):
    if not continuous:
        asset_timeseries = df_to_tuple_dictionary(df)
        op1 = (len(asset_timeseries.keys()) - 2) * entropy(np.column_stack(list(asset_timeseries.values())))
        op2 = np.sum([entropy(asset_timeseries[asset]) - entropy(np.column_stack([v for k, v in asset_timeseries.items() if k != asset])) for asset in asset_timeseries.keys()])
        return op1 + op2
    else:
        if isinstance(df, dict):
            df = pd.DataFrame.from_dict(df)
        return kraskov_jidt.calc_oinfo_kraskov(df)
    # return total_correlation(asset_timeseries) - dual_correlation(asset_timeseries)

def check_validity_of_Oinfo(df):  
    assets_timeseries = df_to_tuple_dictionary(df)
    # O(X, Y) = 0 for any X and Y
    for asset1, asset2 in list(itertools.combinations(assets_timeseries.keys(), 2)):
        filtered_dict = {k: v for k, v in assets_timeseries.items() if k in [asset1, asset2]}
        if Oinfo(filtered_dict) != 0:
            print(f"Oinfo of {asset1}, {asset2} is not equal to 0")

    # O(X, Y, Z) = I(X;Y;Z) for any X, Y, Z
    for asset1, asset2, asset3 in list(itertools.combinations(assets_timeseries.keys(), 3)):
        filtered_dict = {k: v for k, v in assets_timeseries.items() if k in [asset1, asset2, asset3]}
        if round(Oinfo(filtered_dict), 5) != round(MI_three(filtered_dict[asset1], filtered_dict[asset2], filtered_dict[asset3]), 5):
            print(f"Oinfo and three_MI are not equal for {asset1}, {asset2}, {asset3}")

def MI_three(X, Y, Z):
    return MI_timeseries(X, Y) + MI_timeseries(X, Z) - MI_timeseries(X, np.column_stack((Y, Z)))

def Oinfo_bounds(num_variables, cardinality):
    bound = (num_variables - 2) * (np.log(cardinality))
    return (-bound, bound)

# ========================= HELPER FUNCTIONS ====================================
def tuple_if_list(x):
    return tuple(x) if isinstance(x, list) else x

def df_to_tuple_dictionary(df):
    if not isinstance(df, dict):
        return {asset: [tuple_if_list(x) for x in df[asset].values] for asset in df.columns}
    else:
        return df
